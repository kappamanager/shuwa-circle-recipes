"""
activities.xlsx を読んで index.html の ACTIVITIES 配列を書き換える。

xlsx を編集したらこのスクリプトを実行：
  cd 003_site
  python build.py

書き換え対象は index.html の `const ACTIVITIES = [ ... ];` ブロックのみ。
他の HTML/CSS/JS 部分には触れない。
"""

import json
import os
import re
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
HTML_PATH = os.path.join(HERE, "index.html")
XLSX_PATH = os.path.join(HERE, "activities.xlsx")

# init_xlsx.py と同じ列順
COLUMN_ORDER = [
    "id", "title", "catch", "major", "minor", "time", "people", "diff", "prep",
    "purpose", "design", "flow", "schedule_total", "schedule_steps", "tips", "materials",
]

# JS リテラル出力時のフィールド順（index.html の既存スタイルに合わせる）
JS_FIELD_ORDER = [
    "id", "title", "catch", "major", "minor", "time", "people", "diff", "prep",
    "purpose", "design", "flow", "schedule", "tips", "materials",
]


def cell_to_array_or_string(text):
    """セル値が改行を含めば配列、含まなければ string。空なら空文字列"""
    if text is None:
        return ""
    s = str(text).rstrip()
    if not s:
        return ""
    lines = [l.strip() for l in s.split("\n") if l.strip()]
    if len(lines) <= 1:
        return lines[0] if lines else ""
    return lines


def parse_schedule(total, steps_text):
    """schedule_total と schedule_steps セル → schedule オブジェクト or None"""
    total = (total or "").strip() if isinstance(total, str) else (str(total).strip() if total else "")
    steps_text = (steps_text or "").strip() if isinstance(steps_text, str) else ""
    if not total and not steps_text:
        return None

    steps = []
    for line in steps_text.split("\n"):
        line = line.strip()
        if not line:
            continue
        # "5分：ルール説明" or "5分:ルール説明" の形式（半角全角どちらの : も許可）
        m = re.match(r"^(.+?)[：:]\s*(.+)$", line)
        if m:
            minute, label = m.group(1).strip(), m.group(2).strip()
        else:
            # コロンが無ければ全部 label として扱う
            minute, label = "", line
        steps.append({"label": label, "min": minute})

    return {"total": total, "steps": steps}


def js_string(s):
    """Python string → JS double-quoted string。エスケープは \\ と " のみ"""
    if s is None:
        s = ""
    s = str(s)
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    # JS リテラルの中で改行は禁止 → \n に変換（基本的に1行データなので不要だが念のため）
    s = s.replace("\n", "\\n").replace("\r", "")
    return f'"{s}"'


def js_value(v):
    """値を JS リテラル文字列に。
    - None → null
    - bool → true/false
    - int/float → そのまま
    - str → ダブルクォート文字列
    - list → [...]
    - dict → {key:value, ...}（キーは未クォート、JS スタイル）"""
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, str):
        return js_string(v)
    if isinstance(v, list):
        return "[" + ",".join(js_value(x) for x in v) + "]"
    if isinstance(v, dict):
        parts = []
        for k, val in v.items():
            parts.append(f"{k}:{js_value(val)}")
        return "{" + ",".join(parts) + "}"
    raise TypeError(f"Unsupported type: {type(v)}")


def activity_to_js_literal(act):
    """activity dict → JS オブジェクトリテラル文字列"""
    parts = []
    for k in JS_FIELD_ORDER:
        if k not in act:
            continue
        parts.append(f"{k}:{js_value(act[k])}")
    return "{" + ",".join(parts) + "}"


def load_activities_from_xlsx():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["activities"]

    # ヘッダ確認
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(COLUMN_ORDER) + 1)]
    for expected, actual in zip(COLUMN_ORDER, headers):
        if expected != actual:
            raise ValueError(f"列ヘッダ不整合: 期待={expected}, 実際={actual}")

    activities = []
    seen_ids = {}
    warnings = []
    for row in range(2, ws.max_row + 1):
        # id 列が空なら行スキップ
        id_cell = ws.cell(row=row, column=1).value
        if id_cell is None or id_cell == "":
            # id が空でも他のセルに何か入っていれば警告
            other_filled = any(
                ws.cell(row=row, column=c).value not in (None, "")
                for c in range(2, len(COLUMN_ORDER) + 1)
            )
            if other_filled:
                title = ws.cell(row=row, column=2).value or "(タイトル無し)"
                warnings.append(f"  行 {row}: id 空欄のためスキップ（タイトル: {title}）")
            continue

        # id 整数チェック（小数を黙って切り捨てしないよう厳密に判定）
        if isinstance(id_cell, bool):
            raise ValueError(
                f"行 {row}: id に真偽値は使えません（値: {id_cell!r}）"
            )
        if isinstance(id_cell, int):
            rid = id_cell
        elif isinstance(id_cell, float):
            if not id_cell.is_integer():
                raise ValueError(
                    f"行 {row}: id は整数である必要があります（小数値: {id_cell}）"
                )
            rid = int(id_cell)
        elif isinstance(id_cell, str):
            try:
                rid = int(id_cell.strip())
            except ValueError:
                raise ValueError(
                    f"行 {row}: id は整数である必要があります（文字列: {id_cell!r}）"
                )
        else:
            raise ValueError(
                f"行 {row}: id の型が想定外です（値: {id_cell!r}, 型: {type(id_cell).__name__}）"
            )

        # id 重複チェック
        if rid in seen_ids:
            raise ValueError(
                f"行 {row}: id={rid} が重複しています（既出: 行 {seen_ids[rid]}）"
            )
        seen_ids[rid] = row

        row_data = {}
        for col_idx, name in enumerate(COLUMN_ORDER, start=1):
            row_data[name] = ws.cell(row=row, column=col_idx).value

        # schedule の片方だけ入力されているケースを警告
        st = row_data.get("schedule_total")
        ss = row_data.get("schedule_steps")
        if (st and not ss) or (ss and not st):
            warnings.append(
                f"  行 {row} (id={rid}): schedule_total と schedule_steps は"
                f"両方埋めるか両方空にしてください（現状: total={st!r}, steps={'有' if ss else '無'}）"
            )

        sched = parse_schedule(st, ss)

        act = {
            "id": rid,
            "title": (row_data.get("title") or "").strip(),
            "catch": (row_data.get("catch") or "").strip(),
            "major": (row_data.get("major") or "").strip(),
            "minor": (row_data.get("minor") or "").strip(),
            "time": (row_data.get("time") or "").strip(),
            "people": (row_data.get("people") or "").strip(),
            "diff": (row_data.get("diff") or "").strip(),
            "prep": (row_data.get("prep") or "").strip(),
            "purpose": (row_data.get("purpose") or "").strip(),
            "design": cell_to_array_or_string(row_data.get("design")),
            "flow": cell_to_array_or_string(row_data.get("flow")),
            "schedule": sched,
            "tips": cell_to_array_or_string(row_data.get("tips")),
            "materials": (row_data.get("materials") or "").strip(),
        }
        # flow は空でも配列で揃える（既存スタイルに合わせ）
        if isinstance(act["flow"], str):
            act["flow"] = [act["flow"]] if act["flow"] else []
        activities.append(act)

    if warnings:
        print("--- 警告 ---")
        for w in warnings:
            print(w)
        print("")

    return activities


def write_activities_to_html(activities):
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    # 各 activity を JS リテラルに
    body = ",\n  ".join(activity_to_js_literal(a) for a in activities)
    new_block = f"const ACTIVITIES = [\n  {body},\n];"

    # 置換（lambda で渡してバックスラッシュエスケープの再解釈を回避）
    new_html, n = re.subn(
        r"const\s+ACTIVITIES\s*=\s*\[.*?\];",
        lambda _m: new_block,
        html,
        count=1,
        flags=re.DOTALL,
    )
    if n != 1:
        raise RuntimeError("ACTIVITIES ブロックの置換に失敗")

    with open(HTML_PATH, "w", encoding="utf-8", newline="\n") as f:
        f.write(new_html)


def main():
    try:
        activities = load_activities_from_xlsx()
    except PermissionError as e:
        print("--- エラー ---")
        print("activities.xlsx を開けません。")
        print("Excel で activities.xlsx を開いている場合は、保存して閉じてからもう一度実行してください。")
        print(f"詳細: {e}")
        raise SystemExit(1)
    except FileNotFoundError as e:
        print("--- エラー ---")
        print(f"activities.xlsx が見つかりません: {XLSX_PATH}")
        print(f"詳細: {e}")
        raise SystemExit(1)
    except ValueError as e:
        print("--- エラー ---")
        print(f"xlsx の内容に問題があります: {e}")
        raise SystemExit(1)

    print(f"読み込み: {len(activities)} 件")

    try:
        write_activities_to_html(activities)
    except PermissionError as e:
        print("--- エラー ---")
        print("index.html に書き込めません。")
        print("ブラウザ等でファイルがロックされていないか確認してください。")
        print(f"詳細: {e}")
        raise SystemExit(1)
    print(f"書き出し完了: {HTML_PATH}")


if __name__ == "__main__":
    main()
