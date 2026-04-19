"""
index.html の ACTIVITIES 配列を読んで activities.xlsx を生成する初期化スクリプト。
1度だけ実行する。以降は build.py で xlsx → index.html の方向に同期する。

xlsx のフォーマット規則:
  - design / tips: セル内が1行 → string、複数行 → 配列
  - flow:           常に配列。1行 = 1ステップ
  - schedule_total: 空 → schedule:null、値あり → schedule object 生成
  - schedule_steps: 空 OK。各行 "5分：ルール説明" or "5分:ルール説明" 形式

使い方:
  cd 003_site
  python init_xlsx.py
"""

import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
HTML_PATH = os.path.join(HERE, "index.html")
XLSX_PATH = os.path.join(HERE, "activities.xlsx")

# 列定義: (列名, 幅, ラップ表示するか)
COLUMNS = [
    ("id", 6, False),
    ("title", 24, False),
    ("catch", 40, True),
    ("major", 10, False),
    ("minor", 12, False),
    ("time", 10, False),
    ("people", 12, False),
    ("diff", 10, False),
    ("prep", 8, False),
    ("purpose", 50, True),
    ("design", 60, True),
    ("flow", 60, True),
    ("schedule_total", 12, False),
    ("schedule_steps", 50, True),
    ("tips", 60, True),
    ("materials", 40, True),
]


def extract_activities_block(html: str) -> str:
    """index.html から const ACTIVITIES = [ ... ]; のブロック中身を抜き出す"""
    m = re.search(r"const\s+ACTIVITIES\s*=\s*\[(.*?)\];", html, re.DOTALL)
    if not m:
        raise ValueError("ACTIVITIES 配列が見つからない")
    return m.group(1)


def js_to_json(js_text: str) -> str:
    """JS オブジェクトリテラルを JSON 文字列に変換する。
    制限: 値の文字列内に ASCII の : が含まれるケースは想定外（Japanese ：は別文字なのでOK）"""
    # キー名（小文字始まりの英字列）の前にダブルクォートを付ける
    # {id: や ,id: の直後の英字キーが対象
    converted = re.sub(r'([{,])\s*([a-z][a-zA-Z]*)\s*:', r'\1"\2":', js_text)
    # 末尾のカンマを除去
    converted = re.sub(r",\s*$", "", converted.strip())
    return "[" + converted + "]"


def array_or_string_to_cell(value):
    """JSON の string / array → セル内テキスト。配列は改行区切り"""
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(v) for v in value)
    return str(value)


def schedule_to_cells(schedule):
    """schedule オブジェクト → (total文字列, steps改行区切り文字列)"""
    if not schedule:
        return "", ""
    total = schedule.get("total", "") or ""
    steps = schedule.get("steps", []) or []
    lines = []
    for s in steps:
        label = s.get("label", "")
        minute = s.get("min", "")
        lines.append(f"{minute}：{label}")
    return total, "\n".join(lines)


def main():
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    block = extract_activities_block(html)
    json_text = js_to_json(block)
    activities = json.loads(json_text)
    print(f"パース成功: {len(activities)} 件")

    wb = Workbook()
    ws = wb.active
    ws.title = "activities"

    # ヘッダ行
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
    for col_idx, (name, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "B2"  # 1行目固定 + id列固定

    # データ行
    for row_idx, act in enumerate(activities, start=2):
        sched_total, sched_steps = schedule_to_cells(act.get("schedule"))
        row_data = {
            "id": act.get("id"),
            "title": act.get("title", ""),
            "catch": act.get("catch", ""),
            "major": act.get("major", ""),
            "minor": act.get("minor", ""),
            "time": act.get("time", ""),
            "people": act.get("people", ""),
            "diff": act.get("diff", ""),
            "prep": act.get("prep", ""),
            "purpose": act.get("purpose", ""),
            "design": array_or_string_to_cell(act.get("design")),
            "flow": array_or_string_to_cell(act.get("flow")),
            "schedule_total": sched_total,
            "schedule_steps": sched_steps,
            "tips": array_or_string_to_cell(act.get("tips")),
            "materials": act.get("materials", ""),
        }
        for col_idx, (name, _, wrap) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(name, ""))
            if wrap:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    wb.save(XLSX_PATH)
    print(f"出力: {XLSX_PATH}")


if __name__ == "__main__":
    main()
